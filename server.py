from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import openpyxl
from datetime import datetime
import socket
import sys

# Configura a saída do terminal para usar UTF-8 (evita erros com emojis no Windows)
sys.stdout.reconfigure(encoding='utf-8')

app = Flask(__name__, static_folder=".", static_url_path="")
CORS(app)

EXCEL_FILE = "PMMA_2026_Plano_de_Estudos.xlsx"

# Helper para obter o IP local da máquina na rede Wi-Fi
def get_local_ip():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        # Não precisa conectar de verdade, apenas para obter a interface ativa
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

# Auxiliares de limpeza de tipo
def clean_str(val):
    return str(val).strip() if val is not None else ""

def clean_int(val):
    try:
        return int(val) if val is not None else 0
    except (ValueError, TypeError):
        return 0

def clean_float(val):
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    val_str = str(val).split(" ")[0].strip()
    try:
        dt = datetime.strptime(val_str, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        try:
            dt = datetime.strptime(val_str, "%d/%m/%Y")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            return val_str

# Garante que as abas necessárias existam no Excel
def ensure_sheets_exist():
    if not os.path.exists(EXCEL_FILE):
        print(f"Erro: Planilha '{EXCEL_FILE}' não encontrada!")
        return False
    
    wb = openpyxl.load_workbook(EXCEL_FILE)
    modified = False
    
    # 1. Registro_de_Estudos (Histórico)
    if "Registro_de_Estudos" not in wb.sheetnames:
        sheet = wb.create_sheet("Registro_de_Estudos")
        sheet.append([
            "Data", "Matéria", "Assunto", "Tipo", "Tempo (min)", 
            "Qtd Questões", "Acertos", "Erros", "Aproveitamento", "Anotações"
        ])
        modified = True
        print("Aba 'Registro_de_Estudos' criada com sucesso.")
        
    # 2. Caderno_de_Erros
    if "Caderno_de_Erros" not in wb.sheetnames:
        sheet = wb.create_sheet("Caderno_de_Erros")
        sheet.append([
            "Data", "Matéria", "Assunto", "Enunciado/Questão", 
            "Link do TecConcursos", "Minha Resposta", "Explicação/Gabarito"
        ])
        modified = True
        print("Aba 'Caderno_de_Erros' criada com sucesso.")
        
    if modified:
        wb.save(EXCEL_FILE)
        
    return True

# --- ROTAS DE API ---

@app.route("/")
def index():
    return send_from_directory(".", "index.html")

@app.route("/api/data", methods=["GET"])
def get_data():
    if not os.path.exists(EXCEL_FILE):
        return jsonify({"status": "error", "message": f"Planilha '{EXCEL_FILE}' não encontrada."}), 404
        
    try:
        # data_only=True avalia as fórmulas do Excel
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        
        data = {
            "crono": [],
            "edital": [],
            "treinos": [],
            "taf_semanal": [],
            "simulados": [],
            "historico_estudos": [],
            "caderno_erros": []
        }
        
        # 1. Cronograma
        if "Cronograma" in wb.sheetnames:
            sheet = wb["Cronograma"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["crono"].append({
                            "date": format_date(r[0]),
                            "dia": clean_str(r[1]),
                            "semana": clean_str(r[2]),
                            "subject": clean_str(r[3]),
                            "topic": clean_str(r[4]),
                            "type": clean_str(r[5]),
                            "probability": clean_str(r[6]),
                            "hot": "🔥" in clean_str(r[7]),
                            "studied": clean_str(r[8]).lower() in ["sim", "s", "estudado", "concluído", "concluido"],
                            "duration": clean_int(r[9]),
                            "has_questions": clean_str(r[10]).lower() in ["sim", "s"],
                            "questions": clean_int(r[11]),
                            "correct": clean_int(r[12]),
                            "accuracy": clean_float(r[13])
                        })
                        
        # 2. Controle do Edital
        if "Controle_Edital" in wb.sheetnames:
            sheet = wb["Controle_Edital"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["edital"].append({
                            "subject": clean_str(r[0]),
                            "topic": clean_str(r[1]),
                            "probability": clean_str(r[2]),
                            "hot": "🔥" in clean_str(r[3]),
                            "studied": clean_str(r[4]).lower() in ["sim", "s", "estudado", "concluído", "concluido"],
                            "notes": clean_str(r[5]) if len(r) > 5 else ""
                        })
                        
        # 3. Registro de Estudos (Histórico)
        if "Registro_de_Estudos" in wb.sheetnames:
            sheet = wb["Registro_de_Estudos"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["historico_estudos"].append({
                            "date": format_date(r[0]),
                            "subject": clean_str(r[1]),
                            "topic": clean_str(r[2]),
                            "type": clean_str(r[3]),
                            "duration": clean_int(r[4]),
                            "questions": clean_int(r[5]),
                            "correct": clean_int(r[6]),
                            "errors": clean_int(r[7]),
                            "accuracy": clean_float(r[8]),
                            "notes": clean_str(r[9])
                        })
                        
        # 4. Treinos
        if "Treinos" in wb.sheetnames:
            sheet = wb["Treinos"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["treinos"].append({
                            "date": format_date(r[0]),
                            "type": clean_str(r[1]),
                            "duration": clean_int(r[2]),
                            "intensity": clean_int(r[3]),
                            "notes": clean_str(r[4])
                        })
                        
        # 5. TAF Semanal
        if "TAF_Semanal" in wb.sheetnames:
            sheet = wb["TAF_Semanal"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["taf_semanal"].append({
                            "date": format_date(r[0]),
                            "pullups": clean_int(r[1]),
                            "meio_sugado": clean_int(r[2]),
                            "abdominal": clean_int(r[3]),
                            "running": clean_int(r[4]),
                            "status": clean_str(r[5]),
                            "notes": clean_str(r[6])
                        })
                        
        # 6. Simulados Semanais
        if "Simulado_Semanal" in wb.sheetnames:
            sheet = wb["Simulado_Semanal"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["simulados"].append({
                            "date": format_date(r[0]),
                            "number": clean_int(r[1]),
                            "p1_questions": clean_int(r[2]),
                            "p1_correct": clean_int(r[3]),
                            "p2_questions": clean_int(r[4]),
                            "p2_correct": clean_int(r[5]),
                            "score": clean_float(r[6]),
                            "duration": clean_str(r[7]),
                            "notes": clean_str(r[8])
                        })
                        
        # 7. Caderno de Erros
        if "Caderno_de_Erros" in wb.sheetnames:
            sheet = wb["Caderno_de_Erros"]
            rows = list(sheet.iter_rows(values_only=True))
            if len(rows) > 1:
                for r in rows[1:]:
                    if r[0] is not None:
                        data["caderno_erros"].append({
                            "date": format_date(r[0]),
                            "subject": clean_str(r[1]),
                            "topic": clean_str(r[2]),
                            "question": clean_str(r[3]),
                            "link": clean_str(r[4]),
                            "my_answer": clean_str(r[5]),
                            "explanation": clean_str(r[6])
                        })
                        
        # Sincroniza arquivos de backup e consulta estáticos
        try:
            import json
            with open("pmma_data_export.json", "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            import subprocess
            import sys
            subprocess.run([sys.executable, "scratch/generate_crono_html.py"], capture_output=True)
        except Exception as sync_err:
            print(f"Erro ao salvar backups estáticos: {sync_err}")
            
        return jsonify({"status": "success", "data": data})
        
    except Exception as e:
        print(f"Erro ao ler planilha: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao ler planilha: {str(e)}"}), 500

@app.route("/api/study/bulk", methods=["POST"])
def add_study_bulk():
    req_list = request.json
    if not req_list or not isinstance(req_list, list):
        return jsonify({"status": "error", "message": "Corpo da requisição vazio ou formato inválido (esperado array)"}), 400
        
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        
        # Garante que a aba existe
        if "Registro_de_Estudos" not in wb.sheetnames:
            ensure_sheets_exist()
            wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
            
        sheet_reg = wb["Registro_de_Estudos"]
        sheet_edit = wb["Controle_Edital"] if "Controle_Edital" in wb.sheetnames else None
        sheet_crono = wb["Cronograma"] if "Cronograma" in wb.sheetnames else None
        
        for item_data in req_list:
            date_str = item_data.get("date", datetime.today().strftime("%d/%m/%Y"))
            subject = item_data.get("subject", "")
            topic = item_data.get("topic", "")
            study_type = item_data.get("type", "Revisão")
            duration = clean_int(item_data.get("duration", 0))
            questions = clean_int(item_data.get("questions", 0))
            correct = clean_int(item_data.get("correct", 0))
            notes = item_data.get("notes", "Registro em Lote")
            
            errors = max(0, questions - correct)
            accuracy = correct / questions if questions > 0 else 0.0
            
            # 1. Registrar no histórico
            sheet_reg.append([date_str, subject, topic, study_type, duration, questions, correct, errors, accuracy, notes])
            
            # 2. Registrar no Edital
            if sheet_edit:
                updated_edital = False
                for row_idx in range(2, sheet_edit.max_row + 1):
                    sub_val = clean_str(sheet_edit.cell(row=row_idx, column=1).value)
                    top_val = clean_str(sheet_edit.cell(row=row_idx, column=2).value)
                    if sub_val.lower() == subject.lower() and top_val.lower() == topic.lower():
                        if study_type.lower() not in ["simulado", "revisão", "taf"]:
                            sheet_edit.cell(row=row_idx, column=5, value="Sim")
                            if notes:
                                sheet_edit.cell(row=row_idx, column=6, value=notes)
                        updated_edital = True
                        break
                if not updated_edital and subject and topic:
                    if study_type.lower() not in ["simulado", "revisão", "taf"]:
                        sheet_edit.append([subject, topic, "Alta", "Não", "Sim", notes])
                        
            # 3. Registrar no Cronograma
            if sheet_crono:
                updated_crono = False
                for row_idx in range(2, sheet_crono.max_row + 1):
                    sub_val = clean_str(sheet_crono.cell(row=row_idx, column=4).value)
                    top_val = clean_str(sheet_crono.cell(row=row_idx, column=5).value)
                    type_val = clean_str(sheet_crono.cell(row=row_idx, column=6).value)
                    if sub_val.lower() == subject.lower() and top_val.lower() == topic.lower() and type_val.lower() == study_type.lower():
                        sheet_crono.cell(row=row_idx, column=9, value="Sim")
                        sheet_crono.cell(row=row_idx, column=10, value=duration)
                        sheet_crono.cell(row=row_idx, column=11, value="Sim" if questions > 0 else "Não")
                        sheet_crono.cell(row=row_idx, column=12, value=questions)
                        sheet_crono.cell(row=row_idx, column=13, value=correct)
                        sheet_crono.cell(row=row_idx, column=14, value=accuracy)
                        updated_crono = True
                        break
                if not updated_crono and subject and topic:
                    sheet_crono.append([
                        date_str, "Extra", "Semana Extra", subject, topic, study_type, 
                        "Alta", "Não", "Sim", duration, "Sim" if questions > 0 else "Não",
                        questions, correct, accuracy
                    ])
                    
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success", "message": f"{len(req_list)} itens registrados com sucesso em lote!"})
        
    except Exception as e:
        print(f"Erro ao salvar lote: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao salvar lote: {str(e)}"}), 500

@app.route("/api/study", methods=["POST"])
def add_study():
    req = request.json
    if not req:
        return jsonify({"status": "error", "message": "Corpo da requisição vazio"}), 400
        
    date_str = req.get("date", datetime.today().strftime("%d/%m/%Y"))
    subject = req.get("subject", "")
    topic = req.get("topic", "")
    study_type = req.get("type", "Teoria + Questões")
    duration = clean_int(req.get("duration", 0))
    questions = clean_int(req.get("questions", 0))
    correct = clean_int(req.get("correct", 0))
    notes = req.get("notes", "")
    completed = req.get("completed", True)
    
    errors = max(0, questions - correct)
    accuracy = correct / questions if questions > 0 else 0.0
    
    try:
        # Importante: data_only=False preserva as formulas do Excel
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        
        # 1. Registrar na aba Registro_de_Estudos (apenas se completed for True)
        if completed:
            if "Registro_de_Estudos" not in wb.sheetnames:
                ensure_sheets_exist()
                wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
                
            sheet_reg = wb["Registro_de_Estudos"]
            sheet_reg.append([date_str, subject, topic, study_type, duration, questions, correct, errors, accuracy, notes])
        
        # 2. Dar check no Controle_Edital
        if "Controle_Edital" in wb.sheetnames:
            sheet_edit = wb["Controle_Edital"]
            updated_edital = False
            for row_idx in range(2, sheet_edit.max_row + 1):
                sub_val = clean_str(sheet_edit.cell(row=row_idx, column=1).value)
                top_val = clean_str(sheet_edit.cell(row=row_idx, column=2).value)
                if sub_val.lower() == subject.lower() and top_val.lower() == topic.lower():
                    if completed:
                        if study_type.lower() not in ["simulado", "revisão", "taf"]:
                            sheet_edit.cell(row=row_idx, column=5, value="Sim")
                            if notes:
                                sheet_edit.cell(row=row_idx, column=6, value=notes)
                    else:
                        sheet_edit.cell(row=row_idx, column=5, value="Não")
                    updated_edital = True
                    break
            # Caso não ache no edital mas seja uma matéria/tópico válido e completed for True, podemos registrar
            if not updated_edital and subject and topic and completed:
                if study_type.lower() not in ["simulado", "revisão", "taf"]:
                    sheet_edit.append([subject, topic, "Alta", "Não", "Sim", notes])
                
        # 3. Sincronizar com o Cronograma
        if "Cronograma" in wb.sheetnames:
            sheet_crono = wb["Cronograma"]
            updated_crono = False
            
            # Procura um registro existente no Cronograma que tenha a mesma matéria, assunto E TIPO
            for row_idx in range(2, sheet_crono.max_row + 1):
                sub_val = clean_str(sheet_crono.cell(row=row_idx, column=4).value)
                top_val = clean_str(sheet_crono.cell(row=row_idx, column=5).value)
                type_val = clean_str(sheet_crono.cell(row=row_idx, column=6).value)
                
                if sub_val.lower() == subject.lower() and top_val.lower() == topic.lower() and type_val.lower() == study_type.lower():
                    if completed:
                        sheet_crono.cell(row=row_idx, column=9, value="Sim")
                        sheet_crono.cell(row=row_idx, column=10, value=duration)
                        sheet_crono.cell(row=row_idx, column=11, value="Sim" if questions > 0 else "Não")
                        sheet_crono.cell(row=row_idx, column=12, value=questions)
                        sheet_crono.cell(row=row_idx, column=13, value=correct)
                        sheet_crono.cell(row=row_idx, column=14, value=accuracy)
                    else:
                        sheet_crono.cell(row=row_idx, column=9, value="Não")
                        sheet_crono.cell(row=row_idx, column=10, value=None)
                        sheet_crono.cell(row=row_idx, column=11, value="Não")
                        sheet_crono.cell(row=row_idx, column=12, value=None)
                        sheet_crono.cell(row=row_idx, column=13, value=None)
                        sheet_crono.cell(row=row_idx, column=14, value=None)
                    updated_crono = True
                    break
                    
            # Se não achou na grade semanal e completed for True, é um "Estudo Solto". Inserimos no final.
            if not updated_crono and subject and topic and completed:
                sheet_crono.append([
                    date_str, "Extra", "Semana Extra", subject, topic, study_type, 
                    "Alta", "Não", "Sim", duration, "Sim" if questions > 0 else "Não",
                    questions, correct, accuracy
                ])
                
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success", "message": "Estudo registrado com sucesso!"})
        
    except Exception as e:
        print(f"Erro ao salvar estudo: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao salvar estudo: {str(e)}"}), 500

@app.route("/api/review", methods=["POST"])
def add_review():
    req = request.json
    if not req:
        return jsonify({"status": "error", "message": "Corpo da requisição vazio"}), 400
        
    date_str = req.get("date", datetime.today().strftime("%d/%m/%Y"))
    subject = req.get("subject", "")
    topic = req.get("topic", "")
    duration = clean_int(req.get("duration", 0))
    questions = clean_int(req.get("questions", 0))
    correct = clean_int(req.get("correct", 0))
    notes = req.get("notes", "Revisão Espaçada")
    
    errors = max(0, questions - correct)
    accuracy = correct / questions if questions > 0 else 0.0
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        
        # 1. Registrar na aba Registro_de_Estudos
        sheet_reg = wb["Registro_de_Estudos"]
        sheet_reg.append([date_str, subject, topic, "Revisão", duration, questions, correct, errors, accuracy, notes])
        
        # 2. Garante check no cronograma
        if "Cronograma" in wb.sheetnames:
            sheet_crono = wb["Cronograma"]
            for row_idx in range(2, sheet_crono.max_row + 1):
                sub_val = clean_str(sheet_crono.cell(row=row_idx, column=4).value)
                top_val = clean_str(sheet_crono.cell(row=row_idx, column=5).value)
                type_val = clean_str(sheet_crono.cell(row=row_idx, column=6).value)
                if sub_val.lower() == subject.lower() and top_val.lower() == topic.lower() and type_val.lower() == "revisão":
                    sheet_crono.cell(row=row_idx, column=9, value="Sim")
                    break
                    
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success", "message": "Revisão registrada com sucesso!"})
        
    except Exception as e:
        print(f"Erro ao salvar revisão: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao salvar revisão: {str(e)}"}), 500

@app.route("/api/simulado", methods=["POST"])
def add_simulado():
    req = request.json
    if not req:
        return jsonify({"status": "error", "message": "Corpo da requisição vazio"}), 400
        
    date_str = req.get("date", datetime.today().strftime("%d/%m/%Y"))
    number = clean_int(req.get("number", 0))
    p1_q = clean_int(req.get("p1_questions", 0))
    p1_c = clean_int(req.get("p1_correct", 0))
    p2_q = clean_int(req.get("p2_questions", 0))
    p2_c = clean_int(req.get("p2_correct", 0))
    score = clean_float(req.get("score", 0.0))
    duration = req.get("duration", "")
    notes = req.get("notes", "")
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        sheet = wb["Simulado_Semanal"]
        
        # Procura por um simulado existente com o mesmo número
        found_row_idx = None
        for row_idx in range(2, sheet.max_row + 1):
            val = clean_int(sheet.cell(row=row_idx, column=2).value)
            if val == number:
                found_row_idx = row_idx
                break
                
        if found_row_idx:
            # Atualiza a linha existente
            sheet.cell(row=found_row_idx, column=1, value=date_str)
            sheet.cell(row=found_row_idx, column=3, value=p1_q)
            sheet.cell(row=found_row_idx, column=4, value=p1_c)
            sheet.cell(row=found_row_idx, column=5, value=p2_q)
            sheet.cell(row=found_row_idx, column=6, value=p2_c)
            sheet.cell(row=found_row_idx, column=7, value=score)
            sheet.cell(row=found_row_idx, column=8, value=duration)
            sheet.cell(row=found_row_idx, column=9, value=notes)
            msg = "Simulado atualizado com sucesso!"
        else:
            # Insere uma nova linha
            sheet.append([date_str, number, p1_q, p1_c, p2_q, p2_c, score, duration, notes])
            msg = "Simulado registrado com sucesso!"
            
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success", "message": msg})
    except Exception as e:
        print(f"Erro ao salvar simulado: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao salvar simulado: {str(e)}"}), 500

@app.route("/api/taf", methods=["POST"])
def add_taf():
    req = request.json
    if not req:
        return jsonify({"status": "error", "message": "Corpo da requisição vazio"}), 400
        
    date_str = req.get("date", datetime.today().strftime("%d/%m/%Y"))
    pullups = clean_int(req.get("pullups", 0))
    meio_sugado = clean_int(req.get("meio_sugado", 0))
    abdominal = clean_int(req.get("abdominal", 0))
    running = clean_int(req.get("running", 0))
    notes = req.get("notes", "")
    
    # Valida índice do TAF PMMA
    pull_pass = pullups >= 4
    sugado_pass = meio_sugado >= 25
    abd_pass = abdominal >= 35
    run_pass = running >= 2400
    status = "Aprovado" if (pull_pass and sugado_pass and abd_pass and run_pass) else "Pendente"
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        sheet = wb["TAF_Semanal"]
        sheet.append([date_str, pullups, meio_sugado, abdominal, running, status, notes])
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success", "message": "Treino TAF registrado com sucesso!"})
    except Exception as e:
        print(f"Erro ao salvar treino TAF: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao salvar TAF: {str(e)}"}), 500

@app.route("/api/erros", methods=["POST"])
def add_erro():
    req = request.json
    if not req:
        return jsonify({"status": "error", "message": "Corpo da requisição vazio"}), 400
        
    date_str = req.get("date", datetime.today().strftime("%d/%m/%Y"))
    subject = req.get("subject", "")
    topic = req.get("topic", "")
    question = req.get("question", "")
    link = req.get("link", "")
    my_answer = req.get("my_answer", "")
    explanation = req.get("explanation", "")
    
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=False)
        sheet = wb["Caderno_de_Erros"]
        sheet.append([date_str, subject, topic, question, link, my_answer, explanation])
        wb.save(EXCEL_FILE)
        return jsonify({"status": "success", "message": "Erro registrado no Caderno com sucesso!"})
    except Exception as e:
        print(f"Erro ao salvar questão no caderno de erros: {str(e)}")
        return jsonify({"status": "error", "message": f"Erro interno ao salvar erro: {str(e)}"}), 500

if __name__ == "__main__":
    print("=" * 60)
    print("🚀 INICIANDO CENTRAL DE INTELIGÊNCIA QG PMMA 2026")
    print("=" * 60)
    
    ensure_sheets_exist()
    
    local_ip = get_local_ip()
    print(f"\n💻 ACESSO LOCAL: http://localhost:5000")
    print(f"📱 ACESSO NO CELULAR (Wi-Fi): http://{local_ip}:5000")
    print(f"📁 PLANILHA VINCULADA: {os.path.abspath(EXCEL_FILE)}")
    print("=" * 60)
    print("Pressione Ctrl+C para encerrar o servidor.\n")
    
    app.run(host="0.0.0.0", port=5000, debug=False)
