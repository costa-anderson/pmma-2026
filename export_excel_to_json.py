import openpyxl
import json
import os
from datetime import datetime

# Caminhos
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(current_dir, 'Plano_de_Estudos_PMMA_2026.xlsx')
output_path = os.path.join(current_dir, 'pmma_data_export.json')

if not os.path.exists(excel_path):
    # Tenta o nome antigo como fallback
    excel_path = os.path.join(current_dir, 'PMMA_2026_Plano_de_Estudos.xlsx')

if not os.path.exists(excel_path):
    print("Erro: Nenhuma planilha encontrada no diretório!")
    exit(1)

print(f"Carregando planilha: {excel_path}")
wb = openpyxl.load_workbook(excel_path, data_only=True)

export_data = {
    "crono": [],
    "edital": [],
    "treinos": [],
    "taf_semanal": [],
    "simulados": [],
    "historico_estudos": []
}

def clean_str(val):
    return str(val).strip() if val is not None else ""

def clean_int(val):
    try:
        return int(val) if val is not None else 0
    except (ValueError, TypeError):
        return 0

def clean_float(val):
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    val_str = str(val).split(" ")[0].strip()
    try:
        dt = datetime.strptime(val_str, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        try:
            dt = datetime.strptime(val_str, "%d/%m/%Y")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            return val_str

def get_sheet_robust(wb, names):
    for name in names:
        if name in wb.sheetnames:
            return wb[name]
    return None

# 1. Cronograma
crono_sheet = get_sheet_robust(wb, ['Cronograma de Estudos', 'Cronograma', 'Cronograma_de_Estudos'])
if crono_sheet:
    print(f"Exportando aba Cronograma a partir de: {crono_sheet.title}...")
    rows = list(crono_sheet.iter_rows(values_only=True))
    if len(rows) > 0:
        header = [clean_str(x) for x in rows[0] if x is not None]
        is_new_format = "Matéria 1" in header or "Assunto 1" in header or len(rows[0]) <= 11
        
        if is_new_format:
            for r in rows[1:]:
                if len(r) >= 5:
                    date_str = format_date(r[0])
                    dia = clean_str(r[1])
                    semana = clean_str(r[2])
                    is_completed = clean_str(r[9]).lower() in ["concluído", "concluido", "sim", "s"] if len(r) > 9 else False
                    
                    # Matéria 1
                    if r[3] and r[4]:
                        export_data["crono"].append({
                            "date": date_str,
                            "dia": dia,
                            "semana": semana,
                            "subject": clean_str(r[3]),
                            "topic": clean_str(r[4]),
                            "type": "Teoria + Questões",
                            "probability": "Alta",
                            "hot": False,
                            "studied": is_completed,
                            "duration": 0,
                            "questions": 0,
                            "correct": 0,
                            "accuracy": 0.0
                        })
                    
                    # Matéria 2
                    if len(r) > 6 and r[5] and r[6]:
                        export_data["crono"].append({
                            "date": date_str,
                            "dia": dia,
                            "semana": semana,
                            "subject": clean_str(r[5]),
                            "topic": clean_str(r[6]),
                            "type": "Teoria + Questões",
                            "probability": "Alta",
                            "hot": False,
                            "studied": is_completed,
                            "duration": 0,
                            "questions": 0,
                            "correct": 0,
                            "accuracy": 0.0
                        })
        else:
            for r in rows[1:]:
                if len(r) >= 5 and r[3] is not None:
                    export_data["crono"].append({
                        "date": format_date(r[0]),
                        "dia": clean_str(r[1]),
                        "semana": clean_str(r[2]),
                        "subject": clean_str(r[3]),
                        "topic": clean_str(r[4]),
                        "type": clean_str(r[5]) if len(r) > 5 else "",
                        "probability": clean_str(r[6]) if len(r) > 6 else "",
                        "hot": "🔥" in clean_str(r[7]) if len(r) > 7 else False,
                        "studied": clean_str(r[8]).lower() in ["sim", "s", "estudado", "concluído", "concluido"] if len(r) > 8 else False,
                        "duration": clean_int(r[9]) if len(r) > 9 else 0,
                        "has_questions": clean_str(r[10]).lower() in ["sim", "s"] if len(r) > 10 else False,
                        "questions": clean_int(r[11]) if len(r) > 11 else 0,
                        "correct": clean_int(r[12]) if len(r) > 12 else 0,
                        "accuracy": clean_float(r[13]) if len(r) > 13 else 0.0
                    })

# 2. Controle do Edital
edital_sheet = get_sheet_robust(wb, ['Controle do Edital', 'Controle_Edital', 'Edital'])
if edital_sheet:
    print(f"Exportando aba Edital a partir de: {edital_sheet.title}...")
    rows = list(edital_sheet.iter_rows(values_only=True))
    if len(rows) > 0:
        header = [clean_str(x) for x in rows[0] if x is not None]
        is_new_format = len(header) >= 3 and "Estudado" in header[2]
        
        if is_new_format:
            for r in rows[1:]:
                if len(r) >= 2 and r[0] is not None and r[1] is not None:
                    is_studied = clean_str(r[2]).lower() in ["sim", "s", "estudado", "concluído", "concluido"] if len(r) > 2 else False
                    export_data["edital"].append({
                        "subject": clean_str(r[0]),
                        "topic": clean_str(r[1]),
                        "probability": "Alta",
                        "hot": False,
                        "studied": is_studied,
                        "notes": clean_str(r[6]) if len(r) > 6 else ""
                    })
        else:
            for r in rows[1:]:
                if len(r) >= 2 and r[0] is not None and r[1] is not None:
                    export_data["edital"].append({
                        "subject": clean_str(r[0]),
                        "topic": clean_str(r[1]),
                        "probability": clean_str(r[2]) if len(r) > 2 else "",
                        "hot": "🔥" in clean_str(r[3]) if len(r) > 3 else False,
                        "studied": clean_str(r[4]).lower() in ["sim", "s", "estudado", "concluído", "concluido"] if len(r) > 4 else False,
                        "notes": clean_str(r[5]) if len(r) > 5 else ""
                    })

# 3. Registro de Estudos
registro_sheet = get_sheet_robust(wb, ['Registro de Estudos', 'Registro_de_Estudos', 'Registro Estudos'])
if registro_sheet:
    print(f"Exportando aba Registro de Estudos a partir de: {registro_sheet.title}...")
    rows = list(registro_sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        if len(r) >= 3 and r[1] is not None:
            log_item = {
                "date": format_date(r[0]),
                "subject": clean_str(r[1]),
                "topic": clean_str(r[2]),
                "type": clean_str(r[3]) if len(r) > 3 else "",
                "duration": clean_int(r[4]) if len(r) > 4 else 0,
                "questions": clean_int(r[5]) if len(r) > 5 else 0,
                "correct": clean_int(r[6]) if len(r) > 6 else 0,
                "errors": clean_int(r[7]) if len(r) > 7 else 0,
                "notes": clean_str(r[9]) if len(r) > 9 else ""
            }
            export_data["historico_estudos"].append(log_item)
            
            # Mescla no cronograma
            key_subject = log_item["subject"].lower().strip()
            key_topic = log_item["topic"].lower().strip()
            for c in export_data["crono"]:
                if c["subject"].lower().strip() == key_subject and c["topic"].lower().strip() == key_topic:
                    c["duration"] += log_item["duration"]
                    c["questions"] += log_item["questions"]
                    c["correct"] += log_item["correct"]
                    c["accuracy"] = c["correct"] / c["questions"] if c["questions"] > 0 else 0.0

# 4. Treinos e TAF Semanal
treino_sheet = get_sheet_robust(wb, ['Treino do TAF', 'Treinos', 'Treino_do_TAF'])
if treino_sheet:
    print(f"Exportando aba Treino do TAF a partir de: {treino_sheet.title}...")
    rows = list(treino_sheet.iter_rows(values_only=True))
    if len(rows) > 0:
        header = [clean_str(x) for x in rows[0] if x is not None]
        is_new_format = len(header) > 1 and "Exercício" in header[1]
        
        if is_new_format:
            taf_by_date = {}
            for r in rows[1:]:
                date_str = format_date(r[0])
                if not date_str or len(r) < 3:
                    continue
                if date_str not in taf_by_date:
                    taf_by_date[date_str] = {
                        "date": date_str,
                        "pullups": 0,
                        "meio_sugado": 0,
                        "abdominal": 0,
                        "running": 0,
                        "status": "Pendente",
                        "notes": ""
                    }
                ex = clean_str(r[1]).lower()
                val = clean_int(r[2])
                if "barra" in ex:
                    taf_by_date[date_str]["pullups"] = val
                elif "sugado" in ex:
                    taf_by_date[date_str]["meio_sugado"] = val
                elif "abdominal" in ex or "remador" in ex:
                    taf_by_date[date_str]["abdominal"] = val
                elif "corrida" in ex or "12min" in ex:
                    taf_by_date[date_str]["running"] = val
            
            for t in taf_by_date.values():
                pull_pass = t["pullups"] >= 4
                sugado_pass = t["meio_sugado"] >= 25
                abd_pass = t["abdominal"] >= 35
                run_pass = t["running"] >= 2400
                t["status"] = "Aprovado" if (pull_pass and sugado_pass and abd_pass and run_pass) else "Pendente"
                export_data["taf_semanal"].append(t)
                
            for r in rows[1:]:
                if len(r) >= 2 and r[1] is not None:
                    export_data["treinos"].append({
                        "date": format_date(r[0]),
                        "type": clean_str(r[1]),
                        "duration": clean_int(r[2]) if len(r) > 2 else 0,
                        "intensity": clean_int(r[4]) if len(r) > 4 else 0,
                        "notes": clean_str(r[5]) if len(r) > 5 else ""
                    })
        else:
            for r in rows[1:]:
                if len(r) >= 5 and (r[1] is not None or r[2] is not None or r[3] is not None or r[4] is not None):
                    export_data["taf_semanal"].append({
                        "date": format_date(r[0]),
                        "pullups": clean_int(r[1]),
                        "meio_sugado": clean_int(r[2]),
                        "abdominal": clean_int(r[3]),
                        "running": clean_int(r[4]),
                        "status": clean_str(r[5]) if len(r) > 5 else "",
                        "notes": clean_str(r[6]) if len(r) > 6 else ""
                    })

taf_semanal_sheet = get_sheet_robust(wb, ['Simulados do TAF', 'TAF_Semanal'])
if taf_semanal_sheet:
    rows = list(taf_semanal_sheet.iter_rows(values_only=True))
    if len(rows) > 1:
        export_data["taf_semanal"] = [] # Sobrescreve se houver aba simulada dedicada
        for r in rows[1:]:
            if len(r) >= 5 and (r[1] is not None or r[2] is not None or r[3] is not None or r[4] is not None):
                export_data["taf_semanal"].append({
                    "date": format_date(r[0]),
                    "pullups": clean_int(r[1]),
                    "meio_sugado": clean_int(r[2]),
                    "abdominal": clean_int(r[3]),
                    "running": clean_int(r[4]),
                    "status": clean_str(r[5]) if len(r) > 5 else "",
                    "notes": clean_str(r[6]) if len(r) > 6 else ""
                })

# 5. Simulados Cabecalho
sim_sheet = get_sheet_robust(wb, ['Simulados Cabecalho', 'Simulado_Semanal'])
if sim_sheet:
    print(f"Exportando aba Simulados a partir de: {sim_sheet.title}...")
    rows = list(sim_sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        if len(r) >= 2 and r[1] is not None:
            export_data["simulados"].append({
                "date": format_date(r[0]),
                "number": clean_int(r[1]),
                "p1_questions": clean_int(r[2]) if len(r) > 2 else 0,
                "p1_correct": clean_int(r[3]) if len(r) > 3 else 0,
                "p2_questions": clean_int(r[4]) if len(r) > 4 else 0,
                "p2_correct": clean_int(r[5]) if len(r) > 5 else 0,
                "score": clean_float(r[6]) if len(r) > 6 else 0.0,
                "duration": clean_str(r[7]) if len(r) > 7 else "",
                "notes": clean_str(r[8]) if len(r) > 8 else ""
            })

# Salvar JSON
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(export_data, f, ensure_ascii=False, indent=2)

print(f"\nSucesso! {len(export_data['crono'])} itens do Cronograma e {len(export_data['edital'])} do Edital exportados.")
print(f"Arquivo gerado: {output_path}")
