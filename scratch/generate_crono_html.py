import openpyxl
import os
from collections import OrderedDict
import datetime

def generate_static_crono():
    xlsx_path = "PMMA_2026_Plano_de_Estudos.xlsx"
    if not os.path.exists(xlsx_path):
        print(f"Planilha {xlsx_path} não encontrada.")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Cronograma" not in wb.sheetnames:
        print("Aba 'Cronograma' não encontrada.")
        return

    sheet = wb["Cronograma"]
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) <= 1:
        print("Cronograma está vazio.")
        return

    # Agrupar dados por Semana, depois por Dia
    # Usando OrderedDict para preservar a ordem de inserção do Excel
    crono_data = OrderedDict()

    headers = rows[0]
    for r in rows[1:]:
        if r[0] is None:
            continue
        
        # Ler valores
        raw_date = r[0]
        date_str = ""
        if isinstance(raw_date, datetime.datetime):
            date_str = raw_date.strftime("%d/%m/%Y")
        elif raw_date is not None:
            date_str = str(raw_date).split(" ")[0].strip()
            try:
                dt = datetime.datetime.strptime(date_str, "%Y-%m-%d")
                date_str = dt.strftime("%d/%m/%Y")
            except ValueError:
                pass
                
        dia = str(r[1] or "Extra").strip()
        semana = str(r[2] or "Sem Filtro").strip()
        subject = str(r[3] or "").strip()
        topic = str(r[4] or "").strip()
        type_study = str(r[5] or "").strip()
        prob = str(r[6] or "Média").strip()
        hot = "🔥" in str(r[7] or "")
        studied = str(r[8] or "").strip().lower() in ["sim", "s", "estudado", "concluído", "concluido"]
        questions = r[11]
        correct = r[12]

        if semana not in crono_data:
            crono_data[semana] = OrderedDict()
        
        if dia not in crono_data[semana]:
            crono_data[semana][dia] = {
                "date": date_str,
                "topics": []
            }

        crono_data[semana][dia]["topics"].append({
            "subject": subject,
            "topic": topic,
            "type": type_study,
            "prob": prob,
            "hot": hot,
            "studied": studied,
            "questions": questions,
            "correct": correct
        })

    # Gerar HTML
    html_content = """<!DOCTYPE html>
<html lang="pt-BR" data-theme="dark">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Cronograma de Estudos PMMA 2026 - Consulta Estática</title>
    <!-- FontAwesome para ícones -->
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {
            --bg-main: #0b0f19;
            --bg-card: #151b2d;
            --bg-active: #1d263f;
            --text-primary: #f3f4f6;
            --text-secondary: #9ca3af;
            --accent: #2563eb;
            --accent-glow: rgba(37, 99, 235, 0.15);
            --gold: #f59e0b;
            --gold-glow: rgba(245, 158, 11, 0.15);
            --success: #10b981;
            --success-glow: rgba(16, 185, 129, 0.1);
            --border: rgba(255, 255, 255, 0.08);
            --border-active: rgba(37, 99, 235, 0.4);
            --shadow: 0 4px 20px rgba(0, 0, 0, 0.4);
        }

        [data-theme="light"] {
            --bg-main: #f3f4f6;
            --bg-card: #ffffff;
            --bg-active: #f3f4f6;
            --text-primary: #1f2937;
            --text-secondary: #4b5563;
            --accent: #2563eb;
            --accent-glow: rgba(37, 99, 235, 0.1);
            --gold: #d97706;
            --gold-glow: rgba(217, 119, 6, 0.1);
            --success: #059669;
            --success-glow: rgba(5, 150, 105, 0.1);
            --border: rgba(0, 0, 0, 0.08);
            --border-active: rgba(37, 99, 235, 0.3);
            --shadow: 0 4px 20px rgba(0, 0, 0, 0.05);
        }

        * {
            box-sizing: border-box;
            margin: 0;
            padding: 0;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Helvetica, Arial, sans-serif;
        }

        body {
            background-color: var(--bg-main);
            color: var(--text-primary);
            padding: 1.5rem;
            min-height: 100vh;
            transition: background-color 0.3s, color 0.3s;
        }

        header {
            max-width: 1000px;
            margin: 0 auto 1.5rem auto;
            display: flex;
            justify-content: space-between;
            align-items: center;
            border-bottom: 1px solid var(--border);
            padding-bottom: 1rem;
        }

        .header-title h1 {
            font-size: 1.5rem;
            font-weight: 800;
            letter-spacing: -0.5px;
            color: var(--text-primary);
        }
        
        .header-title p {
            font-size: 0.85rem;
            color: var(--text-secondary);
            margin-top: 0.25rem;
        }

        .theme-btn {
            background: var(--bg-card);
            border: 1px solid var(--border);
            color: var(--text-primary);
            padding: 8px 12px;
            border-radius: 8px;
            cursor: pointer;
            font-size: 0.9rem;
            display: flex;
            align-items: center;
            gap: 8px;
            box-shadow: var(--shadow);
            transition: all 0.2s;
        }

        .theme-btn:hover {
            border-color: var(--border-active);
        }

        main {
            max-width: 1000px;
            margin: 0 auto;
        }

        /* Seletor de Semanas */
        .weeks-nav {
            display: flex;
            gap: 0.5rem;
            overflow-x: auto;
            padding-bottom: 0.75rem;
            margin-bottom: 1.5rem;
            scroll-behavior: smooth;
        }

        /* Oculta scrollbar */
        .weeks-nav::-webkit-scrollbar {
            height: 4px;
        }
        .weeks-nav::-webkit-scrollbar-thumb {
            background: var(--border);
            border-radius: 2px;
        }

        .week-btn {
            background: var(--bg-card);
            border: 1px solid var(--border);
            color: var(--text-secondary);
            padding: 8px 16px;
            border-radius: 20px;
            cursor: pointer;
            white-space: nowrap;
            font-size: 0.85rem;
            font-weight: 600;
            transition: all 0.2s;
        }

        .week-btn:hover {
            color: var(--text-primary);
            border-color: var(--border-active);
        }

        .week-btn.active {
            background: var(--accent);
            color: white;
            border-color: var(--accent);
            box-shadow: 0 4px 12px var(--accent-glow);
        }

        /* Container de Semanas */
        .week-container {
            display: none;
            animation: fadeIn 0.3s ease-in-out;
        }

        .week-container.active {
            display: block;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }

        /* Acordeão de Dias */
        .day-group {
            background: var(--bg-card);
            border: 1px solid var(--border);
            border-radius: 12px;
            margin-bottom: 1rem;
            overflow: hidden;
            box-shadow: var(--shadow);
        }

        .day-header {
            padding: 1rem 1.25rem;
            background: rgba(255, 255, 255, 0.01);
            display: flex;
            justify-content: space-between;
            align-items: center;
            cursor: pointer;
            user-select: none;
            transition: background 0.2s;
        }

        .day-header:hover {
            background: rgba(255, 255, 255, 0.03);
        }

        .day-title {
            display: flex;
            align-items: center;
            gap: 0.75rem;
            font-size: 1.05rem;
            font-weight: 700;
        }

        .day-title i {
            color: var(--accent);
        }

        .day-stats {
            display: flex;
            align-items: center;
            gap: 1rem;
            color: var(--text-secondary);
            font-size: 0.8rem;
        }

        .badge-done {
            background: var(--success-glow);
            color: var(--success);
            padding: 2px 8px;
            border-radius: 12px;
            font-weight: 700;
            border: 1px solid rgba(16, 185, 129, 0.2);
        }

        .day-arrow {
            transition: transform 0.2s;
            color: var(--text-secondary);
        }

        .day-group.collapsed .day-arrow {
            transform: rotate(-90deg);
        }

        .day-content {
            padding: 1.25rem;
            border-top: 1px solid var(--border);
            display: flex;
            flex-direction: column;
            gap: 1rem;
        }

        .day-group.collapsed .day-content {
            display: none;
        }

        /* Cartão de Tópico */
        .topic-card {
            background: rgba(255, 255, 255, 0.02);
            border: 1px solid var(--border);
            border-radius: 10px;
            padding: 1rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 1rem;
            transition: transform 0.2s, border-color 0.2s;
        }

        .topic-card:hover {
            border-color: var(--border-active);
        }

        .topic-card.completed {
            border-left: 4px solid var(--success);
            background: rgba(16, 185, 129, 0.02);
        }

        .topic-card.pending {
            border-left: 4px solid var(--text-secondary);
        }

        .topic-info {
            flex: 1;
        }

        .topic-meta {
            display: flex;
            flex-wrap: wrap;
            gap: 0.5rem;
            margin-bottom: 0.5rem;
            align-items: center;
        }

        .subject-badge {
            background: var(--accent-glow);
            color: var(--accent);
            padding: 2px 8px;
            border-radius: 6px;
            font-size: 0.7rem;
            font-weight: 700;
            text-transform: uppercase;
        }

        .prob-badge {
            background: rgba(255, 255, 255, 0.05);
            color: var(--text-secondary);
            padding: 2px 6px;
            border-radius: 4px;
            font-size: 0.7rem;
            font-weight: 600;
        }

        .prob-badge.prob-alta {
            background: var(--gold-glow);
            color: var(--gold);
        }

        .hot-badge {
            color: #ef4444;
            font-size: 0.85rem;
            animation: pulse 1.5s infinite alternate;
        }

        @keyframes pulse {
            from { transform: scale(1); }
            to { transform: scale(1.15); }
        }

        .topic-title {
            font-size: 0.95rem;
            font-weight: 600;
            line-height: 1.4;
            color: var(--text-primary);
        }

        .topic-details {
            font-size: 0.75rem;
            color: var(--text-secondary);
            margin-top: 0.35rem;
            display: flex;
            gap: 0.75rem;
        }

        .status-col {
            display: flex;
            align-items: center;
            justify-content: center;
        }

        .status-icon {
            font-size: 1.25rem;
        }

        .status-icon.done {
            color: var(--success);
        }

        .status-icon.pending {
            color: var(--text-secondary);
            opacity: 0.4;
        }

        /* Responsividade */
        @media (max-width: 768px) {
            body {
                padding: 1rem;
            }
            .topic-card {
                flex-direction: column;
                align-items: flex-start;
                gap: 0.75rem;
            }
            .status-col {
                width: 100%;
                justify-content: flex-end;
                border-top: 1px solid var(--border);
                padding-top: 0.5rem;
            }
        }
    </style>
</head>
<body>
    <header>
        <div class="header-title">
            <h1>QG PMMA 2026</h1>
            <p>Cronograma Geral - Consulta Estática (Fail-Safe)</p>
        </div>
        <button class="theme-btn" id="theme-btn">
            <i class="fa-solid fa-sun"></i> <span>Tema</span>
        </button>
    </header>

    <main>
        <!-- Menu das Semanas -->
        <nav class="weeks-nav">
"""

    # Gerar os botões das semanas
    active_class = "active"
    for w_name in crono_data.keys():
        html_content += f'            <button class="week-btn {active_class}" onclick="selectWeek(\'{w_name}\')" id="btn-{w_name.replace(" ", "-")}">{w_name}</button>\n'
        active_class = ""

    html_content += "        </nav>\n\n"

    # Gerar os containers de cada semana
    active_class = "active"
    for w_name, w_days in crono_data.items():
        html_content += f'        <!-- Container {w_name} -->\n'
        html_content += f'        <div class="week-container {active_class}" id="week-{w_name.replace(" ", "-")}">\n'
        active_class = ""

        # Loop pelos dias da semana
        for d_name, d_val in w_days.items():
            d_date = d_val["date"]
            d_topics = d_val["topics"]
            
            # Contar totais e concluídos
            total_t = len(d_topics)
            done_t = sum(1 for t in d_topics if t["studied"])
            pct_done = int((done_t / total_t) * 100) if total_t > 0 else 0

            badge_html = f'<span class="badge-done">{done_t}/{total_t} Concluído</span>' if pct_done == 100 else f'<span>{done_t}/{total_t} batido</span>'
            
            # Exibir a data se houver
            date_display = f' <span style="font-size: 0.8rem; font-weight: normal; opacity: 0.6; margin-left: 4px;">({d_date})</span>' if d_date else ''

            html_content += f'            <!-- Grupo {d_name} -->\n'
            html_content += f'            <div class="day-group" id="group-{w_name.replace(" ", "-")}-{d_name}">\n'
            html_content += f'                <div class="day-header" onclick="toggleDay(\'{w_name.replace(" ", "-")}-{d_name}\')">\n'
            html_content += f'                    <div class="day-title">\n'
            html_content += f'                        <i class="fa-solid fa-calendar-day"></i>\n'
            html_content += f'                        <span>{d_name}{date_display}</span>\n'
            html_content += f'                    </div>\n'
            html_content += f'                    <div class="day-stats">\n'
            html_content += f'                        {badge_html}\n'
            html_content += f'                        <i class="fa-solid fa-chevron-down day-arrow"></i>\n'
            html_content += f'                    </div>\n'
            html_content += f'                </div>\n'
            html_content += f'                <div class="day-content">\n'

            # Loop pelos tópicos do dia
            for t in d_topics:
                card_class = "completed" if t["studied"] else "pending"
                status_icon = '<i class="fa-solid fa-circle-check status-icon done"></i>' if t["studied"] else '<i class="fa-regular fa-circle status-icon pending"></i>'
                hot_icon = ' <i class="fa-solid fa-fire hot-badge" title="Tema Quente"></i>' if t["hot"] else ''
                prob_class = "prob-alta" if t["prob"].lower() == "alta" else ""
                
                details_str = f'<span><i class="fa-solid fa-tag"></i> {t["type"]}</span>'
                if t["questions"] is not None and t["questions"] > 0:
                    details_str += f' <span><i class="fa-solid fa-clipboard-question"></i> {t["correct"] or 0}/{t["questions"]} Q</span>'

                html_content += f'                    <!-- Card -->\n'
                html_content += f'                    <div class="topic-card {card_class}">\n'
                html_content += f'                        <div class="topic-info">\n'
                html_content += f'                            <div class="topic-meta">\n'
                html_content += f'                                <span class="subject-badge">{t["subject"]}</span>\n'
                html_content += f'                                <span class="prob-badge {prob_class}">Probabilidade: {t["prob"]}</span>\n'
                html_content += f'                                {hot_icon}\n'
                html_content += f'                            </div>\n'
                html_content += f'                            <div class="topic-title">{t["topic"]}</div>\n'
                html_content += f'                            <div class="topic-details">\n'
                html_content += f'                                {details_str}\n'
                html_content += f'                            </div>\n'
                html_content += f'                        </div>\n'
                html_content += f'                        <div class="status-col">\n'
                html_content += f'                            {status_icon}\n'
                html_content += f'                        </div>\n'
                html_content += f'                    </div>\n'

            html_content += f'                </div>\n'
            html_content += f'            </div>\n'

        html_content += f'        </div>\n\n'

    # Adicionar Scripts JS
    html_content += """    </main>

    <script>
        // Troca de Semanas
        function selectWeek(weekName) {
            // Remove active de todos os botões e containers
            document.querySelectorAll('.week-btn').forEach(btn => btn.classList.remove('active'));
            document.querySelectorAll('.week-container').forEach(c => c.classList.remove('active'));

            // Adiciona active no botão clicado e no respectivo container
            const safeId = weekName.replace(/\s+/g, '-');
            const targetBtn = document.getElementById('btn-' + safeId);
            const targetCont = document.getElementById('week-' + safeId);

            if (targetBtn) targetBtn.classList.add('active');
            if (targetCont) targetCont.classList.add('active');
        }

        // Abre/Fecha Acordeão de Dias
        function toggleDay(groupId) {
            const group = document.getElementById('group-' + groupId);
            if (group) {
                group.classList.toggle('collapsed');
            }
        }

        // Alternância de Tema Escuro / Claro
        const themeBtn = document.getElementById('theme-btn');
        themeBtn.addEventListener('click', () => {
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme');
            const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
            html.setAttribute('data-theme', newTheme);
            localStorage.setItem('crono-theme', newTheme);
            
            // Atualiza ícone
            const icon = themeBtn.querySelector('i');
            if (newTheme === 'light') {
                icon.className = 'fa-solid fa-moon';
            } else {
                icon.className = 'fa-solid fa-sun';
            }
        });

        // Carrega tema preferido
        const savedTheme = localStorage.getItem('crono-theme') || 'dark';
        document.documentElement.setAttribute('data-theme', savedTheme);
        const icon = themeBtn.querySelector('i');
        if (savedTheme === 'light') {
            icon.className = 'fa-solid fa-moon';
        } else {
            icon.className = 'fa-solid fa-sun';
        }
    </script>
</body>
</html>
"""

    with open("crono.html", "w", encoding="utf-8") as f:
        f.write(html_content)
    
    print("crono.html gerado com sucesso!")

if __name__ == "__main__":
    generate_static_crono()
