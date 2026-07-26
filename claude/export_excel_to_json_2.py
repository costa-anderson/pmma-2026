import openpyxl
import json
import os
from datetime import datetime

# Caminhos
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(current_dir, '..', 'PMMA_2026_Plano_de_Estudos.xlsx')
output_path = os.path.join(current_dir, 'pmma_data_export_2.json')

if not os.path.exists(excel_path):
    print(f"Erro: Planilha não encontrada em: {excel_path}")
    exit(1)

print(f"Carregando planilha: {excel_path}")
wb = openpyxl.load_workbook(excel_path, data_only=True)

export_data = {
    "crono": [],
    "edital": [],
    "treinos": [],
    "taf_semanal": [],
    "simulados": []
}

def clean_str(val):
    return str(val).strip() if val is not None else ""

def clean_int(val):
    try:
        return int(val) if val is not None else 0
    except (ValueError, TypeError):
        return 0

def clean_float(val):
    try:
        return float(val) if val is not None else 0.0
    except (ValueError, TypeError):
        return 0.0

def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    # Tenta parsear caso seja string
    val_str = str(val).split(" ")[0].strip()
    try:
        # Se for no formato YYYY-MM-DD
        dt = datetime.strptime(val_str, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        try:
            dt = datetime.strptime(val_str, "%d/%m/%Y")
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            return val_str

# 1. Cronograma
print("Exportando aba Cronograma...")
if "Cronograma" in wb.sheetnames:
    sheet = wb["Cronograma"]
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        # Data, Dia, Semana, Matéria, Tema, Tipo, Status Probabilidade, 🔥 Quente, Estudei?, Tempo (min), Fiz Questões?, Qtd Questões, Acertos, Taxa de Acerto
        if len(r) >= 5 and r[3] is not None:
            date_str = format_date(r[0])
            export_data["crono"].append({
                "date": date_str,
                "dia": clean_str(r[1]),
                "semana": clean_str(r[2]),
                "subject": clean_str(r[3]),
                "topic": clean_str(r[4]),
                "type": clean_str(r[5]) if len(r) > 5 else "",
                "probability": clean_str(r[6]) if len(r) > 6 else "",
                "hot": "🔥" in clean_str(r[7]) if len(r) > 7 else False,
                "studied": clean_str(r[8]).lower() in ["sim", "s", "estudado", "concluído", "concluido"] if len(r) > 8 else False,
                "duration": clean_int(r[9]) if len(r) > 9 else 0,
                "has_questions": clean_str(r[10]).lower() in ["sim", "s"] if len(r) > 10 else False,
                "questions": clean_int(r[11]) if len(r) > 11 else 0,
                "correct": clean_int(r[12]) if len(r) > 12 else 0,
                "accuracy": clean_float(r[13]) if len(r) > 13 else 0.0
            })

# 2. Controle_Edital
print("Exportando aba Controle_Edital...")
if "Controle_Edital" in wb.sheetnames:
    sheet = wb["Controle_Edital"]
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        # Matéria, Tópico, Status Probabilidade, 🔥 Tema Quente, Estudado?, Observações
        if len(r) >= 2 and r[0] is not None and r[1] is not None:
            export_data["edital"].append({
                "subject": clean_str(r[0]),
                "topic": clean_str(r[1]),
                "probability": clean_str(r[2]) if len(r) > 2 else "",
                "hot": "🔥" in clean_str(r[3]) if len(r) > 3 else False,
                "studied": clean_str(r[4]).lower() in ["sim", "s", "estudado", "concluído", "concluido"] if len(r) > 4 else False,
                "notes": clean_str(r[5]) if len(r) > 5 else ""
            })

# 3. Treinos
print("Exportando aba Treinos...")
if "Treinos" in wb.sheetnames:
    sheet = wb["Treinos"]
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        # Data, Tipo de Treino, Duração (min), Intensidade (1-5), Observações
        if len(r) >= 2 and r[1] is not None:
            export_data["treinos"].append({
                "date": format_date(r[0]),
                "type": clean_str(r[1]),
                "duration": clean_int(r[2]) if len(r) > 2 else 0,
                "intensity": clean_int(r[3]) if len(r) > 3 else 0,
                "notes": clean_str(r[4]) if len(r) > 4 else ""
            })

# 4. TAF_Semanal
print("Exportando aba TAF_Semanal...")
if "TAF_Semanal" in wb.sheetnames:
    sheet = wb["TAF_Semanal"]
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        # Data, Barra (repetições), Meio Sugado (reps/1min), Abdominal Remador (reps/1min), Corrida 12min (metros), Status Geral, Observações
        if len(r) >= 5 and (r[1] is not None or r[2] is not None or r[3] is not None or r[4] is not None):
            export_data["taf_semanal"].append({
                "date": format_date(r[0]),
                "pullups": clean_int(r[1]),
                "meio_sugado": clean_int(r[2]),
                "abdominal": clean_int(r[3]),
                "running": clean_int(r[4]),
                "status": clean_str(r[5]) if len(r) > 5 else "",
                "notes": clean_str(r[6]) if len(r) > 6 else ""
            })

# 5. Simulado_Semanal
print("Exportando aba Simulado_Semanal...")
if "Simulado_Semanal" in wb.sheetnames:
    sheet = wb["Simulado_Semanal"]
    rows = list(sheet.iter_rows(values_only=True))
    for r in rows[1:]:
        # Data, Nº Simulado, Qtd Questões P1 (Gerais), Acertos P1, Qtd Questões P2 (Específicos), Acertos P2, Nota Final (0-120), Tempo Gasto, Observações
        if len(r) >= 2 and r[1] is not None:
            export_data["simulados"].append({
                "date": format_date(r[0]),
                "number": clean_int(r[1]),
                "p1_questions": clean_int(r[2]) if len(r) > 2 else 0,
                "p1_correct": clean_int(r[3]) if len(r) > 3 else 0,
                "p2_questions": clean_int(r[4]) if len(r) > 4 else 0,
                "p2_correct": clean_int(r[5]) if len(r) > 5 else 0,
                "score": clean_float(r[6]) if len(r) > 6 else 0.0,
                "duration": clean_str(r[7]) if len(r) > 7 else "",
                "notes": clean_str(r[8]) if len(r) > 8 else ""
            })

# Salvar JSON
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(export_data, f, ensure_ascii=False, indent=2)

print(f"\nSucesso! {len(export_data['crono'])} itens do Cronograma e {len(export_data['edital'])} do Edital exportados.")
print(f"Arquivo gerado: {output_path}")
